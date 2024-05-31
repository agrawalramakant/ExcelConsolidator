import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def remove_new_sheet_if_exists(file_path):
    book = load_workbook(file_path)
    if 'New Sheet' in book.sheetnames:
        del book['New Sheet']
    book.save('data/May24/Summary sale May-24.xlsx')


def read_and_aggregate(file_path):
    df = pd.read_excel(file_path, header=7)


    aggregations = {
        'Date': 'first',  # Just take the first date encountered
        'Particulars': 'first',  # Same for particulars
        'GSTIN/UIN': 'first',  # And GSTIN/UIN
        'Vch Type': 'first',  # And Voucher Type
        'Taxable 18% GST': 'sum',
        'Taxable 28% GST': 'sum',
        'Taxable 18% IGST': 'sum',
        'Taxable 28% IGST': 'sum',
        'Integrated Tax Amount': 'sum',
        'Central Tax Amount': 'sum',
        'State Tax Amount': 'sum',
        'Invoice Amount': 'first'  # And Invoice Number
    }

    # Group by 'Vch No.' and apply the defined aggregations
    aggregated_df = df.groupby('Vch No.').agg(aggregations).reset_index()
    aggregated_df['Date'] = pd.to_datetime(aggregated_df['Date']).dt.strftime('%d %b %y')
    return aggregated_df


def write_to_excel(file_path, aggregated_df, sheet_name):
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        aggregated_df.to_excel(writer, sheet_name=sheet_name, index=False)


def adjust_column_width(file_path, sheet_name):
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    book.save(file_path)


if __name__ == '__main__':
    file_path = 'data/May24/Summary sale May-24.xlsx'
    sheet_name = 'Summary'
    remove_new_sheet_if_exists(file_path)
    df = read_and_aggregate(file_path)
    write_to_excel(file_path, df, sheet_name)
    adjust_column_width(file_path, sheet_name)

