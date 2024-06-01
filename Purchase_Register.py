import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def remove_new_sheet_if_exists(file_path, sheet_name):
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    book.save(file_path)

def read_and_aggregate(file_path):
    df = pd.read_excel(file_path, header=2)

    # Create new columns initialized to 0
    df['IGST 18%'] = 0.0
    df['IGST 28%'] = 0.0
    df['GST 18%'] = 0.0
    df['GST 28%'] = 0.0
    df['Date'] = ''

    # Populate new columns based on RATE
    for index, row in df.iterrows():
        if row['RATE'] == 18:
            df.at[index, 'IGST 18%'] = row['IGST']
            df.at[index, 'GST 18%'] = row['SGST']
        elif row['RATE'] == 28:
            df.at[index, 'IGST 28%'] = row['IGST']
            df.at[index, 'GST 28%'] = row['SGST']

    # Group by 'Invoice No' and sum the values to ensure one row per 'Invoice No'
    df_grouped = df.groupby('Invoice No').agg({
        'Date': 'first',  
        'Supplier Name': 'first',
        'GSTIN': 'first', 
        'Invoice Date': 'first',
        'Invoice Value': 'first',  # Assuming Invoice Value remains constant for each Invoice No
        'Taxable Value': 'sum',    # Adjust according to your needs
        'IGST 18%': 'sum',
        'IGST 28%': 'sum',
        'GST 18%': 'sum',
        'GST 28%': 'sum'
    }).reset_index()

    # # Format monetary values to two decimal places
    # df_grouped[['Invoice Value', 'Taxable Value', 'IGST 18%', 'IGST 28%', 'GST 18%', 'GST 28%']] = \
    # df_grouped[['Invoice Value', 'Taxable Value', 'IGST 18%', 'IGST 28%', 'GST 18%', 'GST 28%']].applymap(lambda x: float(f"{x:.2f}"))

    return df_grouped

def write_to_excel(file_path, aggregated_df, sheet_name):
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        aggregated_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=7)

def adjust_column_width(file_path, sheet_name):
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    book.save(file_path)


def add_decimals_to_invoice_value(file_path, sheet_name):
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # Identify columns that contain monetary values by header names
    monetary_columns = [cell.column_letter for cell in sheet[8] if 'Value' in cell.value or 'GST' in cell.value]

    # Set number format for monetary columns
    for col in monetary_columns:
        for cell in sheet[col]:
            if cell.row > 8:  # Only format cells below the header row
                cell.number_format = '#,##,##0.00'

    book.save(file_path)


if __name__ == '__main__':
    file_path = 'data/May24/JJM Purchase Sheet.xlsx'
    sheet_name = 'Purchase Summary'
    remove_new_sheet_if_exists(file_path, sheet_name)
    df = read_and_aggregate(file_path)
    write_to_excel(file_path, df, sheet_name)
    adjust_column_width(file_path, sheet_name)
    add_decimals_to_invoice_value(file_path, sheet_name)